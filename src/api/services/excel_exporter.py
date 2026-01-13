"""
Excel Exporter - Generate Excel files for expenses and reports.

P3 Task: Eksport do Excel
Based on: todo/05-br-priority-roadmap.md

Generates:
- Expense lists with B+R qualification
- Monthly/yearly reports
- Summary statistics
"""

from typing import Optional, Dict, List, Any
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import structlog

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = structlog.get_logger()


class ExcelExporter:
    """
    Excel file generator for B+R documentation.
    
    Creates formatted Excel files with Polish headers.
    """
    
    # Column definitions for expenses
    EXPENSE_COLUMNS = [
        ("Numer faktury", "invoice_number", 20),
        ("Data faktury", "invoice_date", 12),
        ("Dostawca", "vendor_name", 30),
        ("NIP dostawcy", "vendor_nip", 12),
        ("Netto", "net_amount", 12),
        ("VAT", "vat_amount", 10),
        ("Brutto", "gross_amount", 12),
        ("Waluta", "currency", 8),
        ("Kategoria", "expense_category", 20),
        ("B+R", "br_qualified", 6),
        ("Kategoria B+R", "br_category", 15),
        ("Uzasadnienie B+R", "br_qualification_reason", 40),
        ("IP Box", "ip_qualified", 8),
        ("Nexus", "nexus_category", 10),
        ("Status", "status", 12),
    ]
    
    # Header style
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
    HEADER_FONT = Font(bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
    
    # B+R qualified style
    BR_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if OPENPYXL_AVAILABLE else None
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not installed, Excel export disabled")
    
    def export_expenses(
        self,
        expenses: List[Dict[str, Any]],
        title: str = "Wydatki B+R",
        year: Optional[int] = None,
        month: Optional[int] = None
    ) -> bytes:
        """
        Export expenses to Excel file.
        
        Args:
            expenses: List of expense dictionaries
            title: Sheet title
            year: Optional year for filename
            month: Optional month for filename
            
        Returns:
            Excel file as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit
        
        # Add header row
        for col, (header, _, width) in enumerate(self.EXPENSE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add data rows
        for row_num, expense in enumerate(expenses, 2):
            for col, (_, field, _) in enumerate(self.EXPENSE_COLUMNS, 1):
                value = expense.get(field)
                
                # Format specific fields
                if field in ("br_qualified", "ip_qualified"):
                    value = "Tak" if value else "Nie"
                elif field in ("net_amount", "vat_amount", "gross_amount"):
                    value = float(value) if value else 0
                elif field == "invoice_date" and value:
                    if isinstance(value, str):
                        value = value[:10]
                
                cell = ws.cell(row=row_num, column=col, value=value)
                
                # Highlight B+R qualified rows
                if expense.get("br_qualified"):
                    cell.fill = self.BR_FILL
        
        # Add summary section
        summary_row = len(expenses) + 3
        ws.cell(row=summary_row, column=1, value="PODSUMOWANIE").font = Font(bold=True)
        
        total_gross = sum(float(e.get("gross_amount", 0) or 0) for e in expenses)
        br_gross = sum(float(e.get("gross_amount", 0) or 0) for e in expenses if e.get("br_qualified"))
        br_count = sum(1 for e in expenses if e.get("br_qualified"))
        
        ws.cell(row=summary_row + 1, column=1, value="Liczba wydatków:")
        ws.cell(row=summary_row + 1, column=2, value=len(expenses))
        
        ws.cell(row=summary_row + 2, column=1, value="Wydatki B+R:")
        ws.cell(row=summary_row + 2, column=2, value=br_count)
        
        ws.cell(row=summary_row + 3, column=1, value="Suma brutto:")
        ws.cell(row=summary_row + 3, column=2, value=total_gross)
        ws.cell(row=summary_row + 3, column=2).number_format = '#,##0.00 zł'
        
        ws.cell(row=summary_row + 4, column=1, value="Suma B+R brutto:")
        ws.cell(row=summary_row + 4, column=2, value=br_gross)
        ws.cell(row=summary_row + 4, column=2).number_format = '#,##0.00 zł'
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.EXPENSE_COLUMNS))}{len(expenses) + 1}"
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Excel export completed", expenses=len(expenses), br_count=br_count)
        
        return output.getvalue()
    
    def export_monthly_report(
        self,
        expenses: List[Dict[str, Any]],
        revenues: List[Dict[str, Any]],
        year: int,
        month: int,
        company_name: str,
        project_name: str
    ) -> bytes:
        """
        Export monthly B+R report to Excel.
        
        Creates multiple sheets with expenses, revenues, and summary.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export")
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie"
        self._create_summary_sheet(ws_summary, expenses, revenues, year, month, company_name, project_name)
        
        # Expenses sheet
        ws_expenses = wb.create_sheet("Wydatki")
        self._create_expenses_sheet(ws_expenses, expenses)
        
        # Revenues sheet (if any)
        if revenues:
            ws_revenues = wb.create_sheet("Przychody")
            self._create_revenues_sheet(ws_revenues, revenues)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Monthly report exported", year=year, month=month)
        
        return output.getvalue()
    
    def _create_summary_sheet(
        self,
        ws,
        expenses: List[Dict[str, Any]],
        revenues: List[Dict[str, Any]],
        year: int,
        month: int,
        company_name: str,
        project_name: str
    ):
        """Create summary sheet."""
        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Raport B+R - {month:02d}/{year}"
        ws["A1"].font = Font(bold=True, size=16)
        
        # Company info
        ws["A3"] = "Firma:"
        ws["B3"] = company_name
        ws["A4"] = "Projekt:"
        ws["B4"] = project_name
        ws["A5"] = "Okres:"
        ws["B5"] = f"{month:02d}/{year}"
        
        # Expense summary
        ws["A7"] = "WYDATKI"
        ws["A7"].font = Font(bold=True)
        
        total_expenses = len(expenses)
        br_expenses = sum(1 for e in expenses if e.get("br_qualified"))
        ip_expenses = sum(1 for e in expenses if e.get("ip_qualified"))
        
        total_gross = sum(float(e.get("gross_amount", 0) or 0) for e in expenses)
        br_gross = sum(float(e.get("gross_amount", 0) or 0) for e in expenses if e.get("br_qualified"))
        ip_gross = sum(float(e.get("gross_amount", 0) or 0) for e in expenses if e.get("ip_qualified"))
        
        ws["A8"] = "Liczba wydatków:"
        ws["B8"] = total_expenses
        ws["A9"] = "Wydatki B+R:"
        ws["B9"] = br_expenses
        ws["A10"] = "Wydatki IP Box:"
        ws["B10"] = ip_expenses
        ws["A11"] = "Suma brutto:"
        ws["B11"] = total_gross
        ws["B11"].number_format = '#,##0.00 zł'
        ws["A12"] = "Suma B+R:"
        ws["B12"] = br_gross
        ws["B12"].number_format = '#,##0.00 zł'
        ws["A13"] = "Suma IP Box:"
        ws["B13"] = ip_gross
        ws["B13"].number_format = '#,##0.00 zł'
        
        # Revenue summary
        if revenues:
            ws["A15"] = "PRZYCHODY"
            ws["A15"].font = Font(bold=True)
            
            total_revenues = len(revenues)
            total_revenue_amount = sum(float(r.get("gross_amount", 0) or 0) for r in revenues)
            ip_revenue = sum(float(r.get("gross_amount", 0) or 0) for r in revenues if r.get("ip_qualified"))
            
            ws["A16"] = "Liczba przychodów:"
            ws["B16"] = total_revenues
            ws["A17"] = "Suma przychodów:"
            ws["B17"] = total_revenue_amount
            ws["B17"].number_format = '#,##0.00 zł'
            ws["A18"] = "Przychody IP:"
            ws["B18"] = ip_revenue
            ws["B18"].number_format = '#,##0.00 zł'
        
        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
    
    def _create_expenses_sheet(self, ws, expenses: List[Dict[str, Any]]):
        """Create expenses sheet."""
        # Add header row
        for col, (header, _, width) in enumerate(self.EXPENSE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add data rows
        for row_num, expense in enumerate(expenses, 2):
            for col, (_, field, _) in enumerate(self.EXPENSE_COLUMNS, 1):
                value = expense.get(field)
                if field in ("br_qualified", "ip_qualified"):
                    value = "Tak" if value else "Nie"
                elif field in ("net_amount", "vat_amount", "gross_amount"):
                    value = float(value) if value else 0
                ws.cell(row=row_num, column=col, value=value)
        
        ws.freeze_panes = "A2"
    
    def _create_revenues_sheet(self, ws, revenues: List[Dict[str, Any]]):
        """Create revenues sheet."""
        columns = [
            ("Numer faktury", "invoice_number", 20),
            ("Data", "invoice_date", 12),
            ("Klient", "client_name", 30),
            ("NIP klienta", "client_nip", 12),
            ("Netto", "net_amount", 12),
            ("VAT", "vat_amount", 10),
            ("Brutto", "gross_amount", 12),
            ("IP Box", "ip_qualified", 8),
            ("Typ IP", "ip_type", 15),
        ]
        
        # Header row
        for col, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Data rows
        for row_num, revenue in enumerate(revenues, 2):
            for col, (_, field, _) in enumerate(columns, 1):
                value = revenue.get(field)
                if field == "ip_qualified":
                    value = "Tak" if value else "Nie"
                elif field in ("net_amount", "vat_amount", "gross_amount"):
                    value = float(value) if value else 0
                ws.cell(row=row_num, column=col, value=value)
        
        ws.freeze_panes = "A2"


def get_excel_exporter() -> ExcelExporter:
    """Get Excel exporter instance."""
    return ExcelExporter()
